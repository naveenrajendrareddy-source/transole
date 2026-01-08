# clientdoc/views.py

from django.template.loader import render_to_string
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib import messages
from django.db import transaction
from django.core.paginator import Paginator
from django.conf import settings
from django.urls import reverse
from .models import SalesInvoice, InvoiceItem, Item, StoreLocation, DeliveryChallan, TransportCharges, ConfirmationDocument, PackedImage, OurCompanyProfile, ActivityLog, Buyer, BulkInvoiceUpload, ItemCategory
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from .forms import InvoiceForm, DeliveryChallanForm, TransportChargesForm, ConfirmationDocumentForm, PackedImageFormSet, ItemForm, StoreLocationForm, BuyerForm, InvoiceItemFormSet
import json
from .pdf_generator import generate_invoice_pdf, generate_dc_pdf, generate_transport_pdf
import logging
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.utils import ImageReader
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as PlatypusImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import mm
from PyPDF2 import PdfMerger, PdfReader
import os

logger = logging.getLogger(__name__)

# --- PDF GENERATION HELPERS ---

def generate_packed_images_pdf(confirmation):
    """Generates a PDF page for packed images."""
    images = confirmation.packedimage_set.all()
    if not images.exists():
        return None 

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    margin = 50
    img_width = width - (2 * margin) 
    img_height = 250 
    spacing = 20
    y = height - margin
    
    c.setFont("Helvetica-Bold", 16)
    c.drawString(margin, y, "Packed Goods Images")
    y -= 40
    
    # Track unique image paths to prevent duplicates
    seen_images = set()
    
    for image_obj in images:
        # Skip duplicate images based on file path
        try:
            img_path = image_obj.image.path
            
            # Check if we've already processed this image
            if img_path in seen_images:
                continue
            
            seen_images.add(img_path)
            
        except Exception:
            # If path access fails, skip this image
            continue
        
        if y < margin + img_height + spacing:
            c.showPage()
            y = height - margin - 20 
            
        try:
            img = ImageReader(img_path) 
            
            aspect = img.getSize()[1] / img.getSize()[0]
            current_img_height = img_width * aspect
            
            if current_img_height > img_height:
                current_img_height = img_height

            c.drawImage(img, margin, y - current_img_height, width=img_width, height=current_img_height)
            
            c.setFont("Helvetica", 10)
            notes_y = y - current_img_height - 10
            c.drawString(margin, notes_y, f"Notes: {image_obj.notes or 'N/A'}")

            y -= (current_img_height + spacing + 20) 

        except Exception as e:
            logger.error(f"Error drawing image {image_obj.id} to PDF: {e}")
            c.setFont("Helvetica-Bold", 12)
            c.drawString(margin, y, f"Error loading image {image_obj.id}: {e}")
            y -= 30
            
    c.save()
    buffer.seek(0)
    return buffer


# --- 1. DASHBOARD & LIST VIEWS (FIX 3: Corrected List Views) ---

def dashboard(request):
    """Shows system overview and recent activity (recent invoices)."""
    invoices = SalesInvoice.objects.all().select_related('location').order_by('-date')[:10]
    total_invoices = SalesInvoice.objects.count()
    total_finalized = SalesInvoice.objects.filter(status='FIN').count()
    
    context = {
        'invoices': invoices,
        'total_invoices': total_invoices,
        'total_finalized': total_finalized
    }
    
    recent_logs = ActivityLog.objects.order_by('-timestamp')[:10]
    
    context['recent_logs'] = recent_logs
    return render(request, 'clientdoc/dashboard.html', context)

def item_detail(request, item_id):
    """Detail view for a single item."""
    item = get_object_or_404(Item, id=item_id)
    return render(request, 'clientdoc/item_detail.html', {'item': item})

def log_activity(action, details=""):
    ActivityLog.objects.create(action=action, details=details)

def get_filtered_queryset(model_class, request, search_fields):
    """Helper to filter and sort querysets."""
    queryset = model_class.objects.all().select_related('invoice') if model_class != SalesInvoice and hasattr(model_class, 'invoice') else model_class.objects.all()
    if model_class == SalesInvoice:
        queryset = queryset.select_related('location')
    elif hasattr(model_class, 'invoice'):
         queryset = queryset.filter(invoice__is_deleted=False)
        
    # Search
    query = request.GET.get('q')
    if query:
        from django.db.models import Q
        q_objects = Q()
        for field in search_fields:
            q_objects |= Q(**{field + '__icontains': query})
        queryset = queryset.filter(q_objects)
    
    # Sort
    # Sort
    sort_by = request.GET.get('sort')
    
    # Determine default sort if not provided
    if not sort_by:
        if hasattr(model_class, 'date'):
            sort_by = '-date'
        elif hasattr(model_class, 'created_at'):
            sort_by = '-created_at'
        else:
            sort_by = '-id'

    if sort_by == 'az': 
        if model_class == SalesInvoice:
            sort_by = 'tally_invoice_number'
        elif hasattr(model_class, 'name'):
            sort_by = 'name'
        else:
            sort_by = 'invoice__tally_invoice_number'

    if sort_by == 'za': 
        if model_class == SalesInvoice:
            sort_by = '-tally_invoice_number'
        elif hasattr(model_class, 'name'):
            sort_by = '-name'
        else:
            sort_by = '-invoice__tally_invoice_number'
            
    # Better date sorting using created_at if available and date is not
    if sort_by in ['date', '-date'] and not hasattr(model_class, 'date') and hasattr(model_class, 'created_at'):
        sort_by = sort_by.replace('date', 'created_at')
        
    # Safety check: If trying to sort by date/created_at but model lacks it
    if 'date' in sort_by and not hasattr(model_class, 'date'):
        sort_by = '-id'
    if 'created_at' in sort_by and not hasattr(model_class, 'created_at'):
        sort_by = '-id'
    
    allowed_sorts = [
         'date', '-date', 
         'created_at', '-created_at', 
         'id', '-id', 
         'total', '-total', 
         'status', '-status', 
         'tally_invoice_number', '-tally_invoice_number', 
         'app_invoice_number', '-app_invoice_number',
         'invoice__tally_invoice_number', '-invoice__tally_invoice_number', 
         'invoice__app_invoice_number', '-invoice__app_invoice_number',
         'invoice__date', '-invoice__date',
         'name', '-name'
    ]
                     
    if sort_by in allowed_sorts:
        queryset = queryset.order_by(sort_by)
    else:
        # Default sorts
        if hasattr(model_class, 'date'):
            queryset = queryset.order_by('-date')
        elif hasattr(model_class, 'invoice'):
             queryset = queryset.order_by('-invoice__date')
        else:
             queryset = queryset.order_by('-id')
        
    return queryset

def trash_list(request):
    """View to show deleted items."""
    invoices = SalesInvoice.objects.trash().all()
    locations = StoreLocation.objects.trash().all()
    items = Item.objects.trash().all()
    
    return render(request, 'clientdoc/trash_list.html', {
        'invoices': invoices,
        'locations': locations,
        'items': items,
        'title': 'Trash Bin'
    })

def restore_object(request, model_name, pk):
    """Restores a soft-deleted object."""
    model_map = {
        'invoice': SalesInvoice,
        'location': StoreLocation,
        'item': Item,
        'dc': DeliveryChallan,
        'transport': TransportCharges,
        'confirmation': ConfirmationDocument,
        'buyer': Buyer
    }
    model = model_map.get(model_name)
    if not model:
        messages.error(request, 'Invalid item type.')
        return redirect('clientdoc:trash_list')
        
    obj = get_object_or_404(model.objects.trash(), pk=pk)
    obj.restore()
    log_activity("Restore", f"Restored {model_name} #{pk}")
    messages.success(request, f'{model_name.title()} restored successfully.')
    return redirect('clientdoc:trash_list')

def hard_delete_object(request, model_name, pk):
    """Permanently deletes an object."""
    model_map = {
        'invoice': SalesInvoice,
        'location': StoreLocation,
        'item': Item,
        'dc': DeliveryChallan,
        'transport': TransportCharges,
        'confirmation': ConfirmationDocument,
        'buyer': Buyer
    }
    model = model_map.get(model_name)
    if not model:
        messages.error(request, 'Invalid item type.')
        return redirect('clientdoc:trash_list')
        
    obj = get_object_or_404(model.objects.trash(), pk=pk)
    obj.hard_delete()
    log_activity("Permanent Delete", f"Permanently deleted {model_name} #{pk}")
    messages.warning(request, f'{model_name.title()} permanently deleted.')
    return redirect('clientdoc:trash_list')

def delete_object(request, model_name, pk):
    """Soft deletes an object from list view."""
    model_map = {
        'invoice': SalesInvoice,
        'location': StoreLocation,
        'item': Item,
        'dc': DeliveryChallan,
        'transport': TransportCharges,
        'confirmation': ConfirmationDocument,
        'buyer': Buyer
    }
    model = model_map.get(model_name)
    if not model:
        messages.error(request, 'Invalid item type.')
        return redirect('clientdoc:dashboard')

    obj = get_object_or_404(model, pk=pk)
    obj.delete() # Soft delete
    log_activity("Delete", f"Moved {model_name} #{pk} to trash")
    messages.success(request, f'{model_name.title()} moved to trash.')
    return redirect(request.META.get('HTTP_REFERER', 'clientdoc:dashboard'))

def invoice_list(request):
    search_fields = ['tally_invoice_number', 'app_invoice_number', 'location__name', 'date']
    invoices = get_filtered_queryset(SalesInvoice, request, search_fields)
    
    paginator = Paginator(invoices, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'clientdoc/invoice_list.html', {
        'page_obj': page_obj, 
        'title': 'Sales Invoice List',
        'list_type': 'inv'
    })

def dc_list(request):
    search_fields = ['invoice__tally_invoice_number', 'invoice__app_invoice_number', 'invoice__location__name', 'date']
    challans = get_filtered_queryset(DeliveryChallan, request, search_fields)
    
    paginator = Paginator(challans, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'clientdoc/dc_list.html', {
        'page_obj': page_obj, 
        'title': 'Delivery Challan List',
        'list_type': 'dc'
    })
    
def transport_list(request):
    search_fields = ['invoice__tally_invoice_number', 'invoice__app_invoice_number', 'invoice__location__name', 'date', 'description']
    charges = get_filtered_queryset(TransportCharges, request, search_fields)
    
    paginator = Paginator(charges, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'clientdoc/transport_list.html', {
        'page_obj': page_obj, 
        'title': 'Transport Charges List',
        'list_type': 'trp'
    })

def confirmation_list(request):
    search_fields = ['invoice__tally_invoice_number', 'invoice__app_invoice_number', 'invoice__location__name', 'date']
    docs = get_filtered_queryset(ConfirmationDocument, request, search_fields)
    
    paginator = Paginator(docs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'clientdoc/confirmation_list.html', {
        'page_obj': page_obj, 
        'title': 'Confirmation Document List',
        'list_type': 'cnf'
    })

# --- Utility Functions (Ensure create_item exists) ---

# --- ITEM VIEWS ---

def item_list(request):
    search_fields = ['name', 'description']
    items = get_filtered_queryset(Item, request, search_fields)
    
    paginator = Paginator(items, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'clientdoc/item_list.html', {
        'page_obj': page_obj, 
        'title': 'Item List',
        'list_type': 'item'
    })

def edit_item(request, pk):
    item = get_object_or_404(Item, pk=pk)
    if request.method == 'POST':
        form = ItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            log_activity("Edit Item", f"Updated Item {item.name}")
            messages.success(request, 'Item updated successfully.')
            return redirect('clientdoc:item_list')
    else:
        form = ItemForm(instance=item)
    return render(request, 'clientdoc/form.html', {'form': form, 'title': 'Edit Item'})

def create_item(request):
    if request.method == 'POST':
        form = ItemForm(request.POST) 
        if form.is_valid():
            form.save()
            log_activity("Create Item", f"Created Item {form.instance.name}")
            messages.success(request, 'Item created successfully.')
            return redirect('clientdoc:dashboard')
    else:
        form = ItemForm()
    return render(request, 'clientdoc/form.html', {'form': form, 'title': 'Create Item'})

def create_location(request):
    if request.method == 'POST':
        form = StoreLocationForm(request.POST) 
        if form.is_valid():
            form.save()
            log_activity("Create Location", f"Created Location {form.instance.name}")
            messages.success(request, 'Location created successfully.')
            return redirect('clientdoc:dashboard')
    else:
        form = StoreLocationForm()
    return render(request, 'clientdoc/form.html', {'form': form, 'title': 'Create Store Location'})

def store_location_list(request):
    search_fields = ['name', 'address', 'city', 'gstin', 'site_code']
    locations = get_filtered_queryset(StoreLocation, request, search_fields)
    
    paginator = Paginator(locations, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'clientdoc/store_location_list.html', {
        'page_obj': page_obj, 
        'title': 'Store Client Locations',
        'list_type': 'location'
    })

def edit_location(request, pk):
    location = get_object_or_404(StoreLocation, pk=pk)
    if request.method == 'POST':
        form = StoreLocationForm(request.POST, instance=location)
        if form.is_valid():
            form.save()
            log_activity("Edit Location", f"Updated Location {location.name}")
            messages.success(request, 'Location updated successfully.')
            return redirect('clientdoc:store_location_list')
    else:
        form = StoreLocationForm(instance=location)
    return render(request, 'clientdoc/form.html', {'form': form, 'title': 'Edit Store Location'})

def store_location_detail(request, pk):
    location = get_object_or_404(StoreLocation, pk=pk)
    return render(request, 'clientdoc/store_location_detail.html', {'location': location})

# --- 2. WORKFLOW STEP 1: CREATE INVOICE ITEMS ---

def create_invoice(request):
    """Handles creation of SalesInvoice and multiple InvoiceItem records using FormSets."""
    
    items = Item.objects.all()
    locations = StoreLocation.objects.all()
    buyers = Buyer.objects.all()
    
    # Use the imported FormSet, or create a factory if specific config needed
    # formset = InvoiceItemFormSet(queryset=InvoiceItem.objects.none()) # If usage of imported one

    if request.method == 'POST':
        location_id = request.POST.get('location')
        buyer_id = request.POST.get('buyer')
        tally_invoice_number = request.POST.get('tally_invoice_number')
        date = request.POST.get('date')

        if not location_id:
            messages.error(request, 'Please select a client location.')
            return redirect('clientdoc:create_invoice')

        try:
            with transaction.atomic():
                location = get_object_or_404(StoreLocation, id=location_id)
                buyer = None
                if buyer_id:
                    buyer = get_object_or_404(Buyer, id=buyer_id)
                
                invoice = SalesInvoice.objects.create(location=location, buyer=buyer, status='DRF') 
                if tally_invoice_number: invoice.tally_invoice_number = tally_invoice_number
                if date: invoice.date = date
                invoice.save()
                
                # Bind formset to new invoice
                formset = InvoiceItemFormSet(request.POST, instance=invoice, prefix='invoiceitem_set')
                
                print(f"DEBUG: TOTAL_FORMS: {request.POST.get('invoiceitem_set-TOTAL_FORMS')}")
                if not formset.is_valid():
                    print(f"DEBUG: Formset Errors: {formset.errors}")
                    print(f"DEBUG: NonForm Errors: {formset.non_form_errors()}")

                print(f"DEBUG: TOTAL_FORMS: {request.POST.get('invoiceitem_set-TOTAL_FORMS')}")
                if formset.is_valid():
                    # Standard save handles foreign keys because instance=invoice is passed
                    formset.save()
                    
                    invoice.refresh_from_db()
                    invoice.calculate_total()
                    
                    log_activity("Create Invoice", f"Created Invoice {invoice.id} with {invoice.invoiceitem_set.count()} items")
                    messages.success(request, f'Invoice #{invoice.id} created successfully!')
                    return redirect('clientdoc:edit_invoice', invoice_id=invoice.id)
                else:
                    # Rollback if items are invalid
                    transaction.set_rollback(True)
                    # Show errors
                    if formset.non_form_errors():
                        messages.error(request, f"Formset Error: {formset.non_form_errors()}")
                    for form in formset:
                        for field, errors in form.errors.items():
                             for error in errors:
                                 messages.error(request, f"Item Error ({field}): {error}")
                    messages.error(request, 'Failed to create invoice. Please check item details.')

        except Exception as e:
            logger.error(f"Invoice Create Error: {e}")
            messages.error(request, f"Error creating invoice: {str(e)}")

    else:
        # GET request - Initialize empty formset so we have management form
        # We pass instance=None or a dummy unsaved instance? 
        # inlineformset factory expects instance. SalesInvoice() is fine.
        formset = InvoiceItemFormSet(instance=SalesInvoice(), prefix='invoiceitem_set')
        formset.extra = 0

    return render(request, 'clientdoc/invoice_form.html', {
        'locations': locations,
        'buyers': buyers,
        'items': items,
        'formset': formset,
        'title': 'Create Sales Invoice'
    })


# --- 3. WORKFLOW STEP 2: EDIT INVOICE (TALLY DETAILS) ---
def edit_invoice(request, invoice_id):
    invoice = get_object_or_404(SalesInvoice, id=invoice_id)
    items = Item.objects.all() 
    
    if request.method == 'POST':
        form = InvoiceForm(request.POST, instance=invoice) 
        formset = InvoiceItemFormSet(request.POST, instance=invoice, prefix='invoiceitem_set')
        
        if form.is_valid() and formset.is_valid():
            form.save()
            
            instances = formset.save(commit=False)
            for instance in instances:
                if instance.item_id:
                    instance.invoice = invoice
                    instance.save()
            
            for obj in formset.deleted_objects:
                obj.delete()
                
            invoice.calculate_total()
            log_activity("Edit Invoice", f"Updated Invoice {invoice.tally_invoice_number or invoice.id} details")
            messages.success(request, f'Invoice details updated.')
            
            if request.POST.get('action') == 'save_continue':
                return redirect('clientdoc:edit_dc', invoice_id=invoice.id)
            if request.POST.get('action') == 'save_list':
                return redirect('clientdoc:invoice_list')
            
            return redirect('clientdoc:edit_invoice', invoice_id=invoice.id) 
        else:
             if not form.is_valid():
                 messages.error(request, f"Header Errors: {form.errors}")
             if not formset.is_valid():
                 messages.error(request, f"Item Errors: {formset.errors}")
    else:
        form = InvoiceForm(instance=invoice)
        formset = InvoiceItemFormSet(instance=invoice, prefix='invoiceitem_set')
    
    invoice.refresh_from_db()
    next_url = reverse('clientdoc:edit_dc', kwargs={'invoice_id': invoice.id})

    return render(request, 'clientdoc/edit_tally_details.html', {
        'form': form,
        'formset': formset,
        'items': items,
        'invoice': invoice, # Fixed
        'title': f'Edit Tally Details for Invoice #{invoice.id}',
        'next_url': next_url, 
        'current_step': 1,
        'progress_percentage': 25,
    })


# --- 4. WORKFLOW STEP 3: EDIT DELIVERY CHALLAN (DC) ---
def edit_dc(request, invoice_id):
    invoice = get_object_or_404(SalesInvoice, id=invoice_id)
    dc, created = DeliveryChallan.objects.get_or_create(invoice=invoice)

    if request.method == 'POST':
        form = DeliveryChallanForm(request.POST, instance=dc)
        if form.is_valid():
            form.save()
            log_activity("Edit DC", f"Updated DC for Invoice {invoice.id}")
            
            if invoice.status == 'DRF':
                invoice.status = 'DC'
                invoice.save()
                
            # FIX 2: Redirect to the DC List after edit
            messages.success(request, 'Delivery Challan updated.')
            
            if request.POST.get('action') == 'save_continue':
                return redirect('clientdoc:edit_transport', invoice_id=invoice.id)
            if request.POST.get('action') == 'save_list':
                return redirect('clientdoc:dc_list')
                
            return redirect('clientdoc:edit_dc', invoice_id=invoice.id) 
    else:
        form = DeliveryChallanForm(instance=dc)
    
    next_url = reverse('clientdoc:edit_transport', kwargs={'invoice_id': invoice.id})
    prev_url = reverse('clientdoc:edit_invoice', kwargs={'invoice_id': invoice.id})

    return render(request, 'clientdoc/form.html', {
        'form': form,
        'title': f'Delivery Challan - Invoice {invoice.tally_invoice_number or invoice.id}',
        'next_url': next_url,
        'prev_url': prev_url, 
        'current_step': 2,
        'progress_percentage': 50,
    })


# --- 5. WORKFLOW STEP 4: EDIT TRANSPORT CHARGES ---
def edit_transport(request, invoice_id):
    invoice = get_object_or_404(SalesInvoice, id=invoice_id)
    transport, created = TransportCharges.objects.get_or_create(invoice=invoice)
        
    if invoice.status not in ['DC', 'TRP', 'FIN']:
        messages.error(request, 'You must complete the Delivery Challan first.')
        return redirect('clientdoc:dashboard')
        
    if request.method == 'POST':
        form = TransportChargesForm(request.POST, instance=transport)
        if form.is_valid():
            form.save()
            log_activity("Edit Transport", f"Updated Transport Charges for Invoice {invoice.id}")
            
            if invoice.status == 'DC':
                invoice.status = 'TRP'
                invoice.save()
                
            # FIX 2: Redirect to the Transport Charges List after edit
            messages.success(request, 'Transport charges updated.')
            
            if request.POST.get('action') == 'save_continue':
                return redirect('clientdoc:create_confirmation', invoice_id=invoice.id)
            if request.POST.get('action') == 'save_list':
                return redirect('clientdoc:transport_list')
                
            return redirect('clientdoc:edit_transport', invoice_id=invoice.id) 
    else:
        form = TransportChargesForm(instance=transport)
    
    next_url = reverse('clientdoc:create_confirmation', kwargs={'invoice_id': invoice.id})
    prev_url = reverse('clientdoc:edit_dc', kwargs={'invoice_id': invoice.id})

    return render(request, 'clientdoc/form.html', {
        'form': form,
        'title': f'Transport Charges - Invoice {invoice.tally_invoice_number or invoice.id}',
        'next_url': next_url,
        'prev_url': prev_url, 
        'current_step': 3,
        'progress_percentage': 75,
    })


# --- 6. WORKFLOW STEP 5: CONFIRMATION & PDF GENERATION (FIX 1: Robust Merging) ---

def create_confirmation(request, invoice_id):
    invoice = get_object_or_404(SalesInvoice, id=invoice_id)
    confirmation, created = ConfirmationDocument.objects.get_or_create(invoice=invoice)
    company_profile = OurCompanyProfile.objects.first() 
    
    if invoice.status not in ['TRP', 'FIN']:
        messages.error(request, 'Cannot access Confirmation Document yet. Please log Transport Charges first.')
        return redirect('clientdoc:dashboard')
    
    # File deletion logic (kept short for brevity)
    if request.method == 'POST':
        if 'delete_po' in request.POST and confirmation.po_file:
            confirmation.po_file.delete(save=False)
            confirmation.po_file = None
            confirmation.save()
            messages.success(request, 'Purchase Order file removed.')
            return redirect('clientdoc:create_confirmation', invoice_id=invoice_id)

        if 'delete_email' in request.POST and confirmation.approval_email_file:
            confirmation.approval_email_file.delete(save=False)
            confirmation.approval_email_file = None
            confirmation.save()
            messages.success(request, 'Approval Email file removed.')
            return redirect('clientdoc:create_confirmation', invoice_id=invoice_id)
    
    has_po = bool(confirmation.po_file)
    has_email = bool(confirmation.approval_email_file)

    if request.method == 'POST':
        form = ConfirmationDocumentForm(request.POST, request.FILES, instance=confirmation)
        image_formset = PackedImageFormSet(request.POST, request.FILES, instance=confirmation)
        
        if form.is_valid() and image_formset.is_valid():
            confirmation = form.save()
            image_formset.save()
            
            if 'save_notes' in request.POST:
                messages.success(request, 'Files and image notes saved successfully.')
                return redirect('clientdoc:create_confirmation', invoice_id=invoice_id)

            # --- REDIRECT TO CHECKLIST INSTEAD OF AUTO FINALIZE ---
            messages.success(request, 'Files and image notes saved successfully. Please review and finalize.')
            return redirect('clientdoc:create_confirmation', invoice_id=invoice_id)
    
    else:
        form = ConfirmationDocumentForm(instance=confirmation)
        image_formset = PackedImageFormSet(instance=confirmation)
    
    prev_url = reverse('clientdoc:edit_transport', kwargs={'invoice_id': invoice.id})
    packed_images_list = confirmation.packedimage_set.all()

    # Prepare available files for Checklist
    available_files = [
        {'id': 'invoice', 'name': 'Tax Invoice (Auto-Generated)', 'required': False},
    ]
    if hasattr(invoice, 'deliverychallan'):
        available_files.append({'id': 'dc', 'name': 'Delivery Challan (Auto-Generated)', 'required': False})
        
    if hasattr(invoice, 'transportcharges'):
        available_files.append({'id': 'transport', 'name': 'Transport Charges (Auto-Generated)', 'required': False})
    if has_po:
        available_files.append({'id': 'po', 'name': 'PO Copy (Uploaded)', 'required': False})
    if has_email:
        available_files.append({'id': 'email', 'name': 'Approval Email (Uploaded)', 'required': False})
    
    # Images are always last usually, but let's allow them in list if we want to be fancy, 
    # but for now images are appended at end in PDF gen logic typically. 
    # Let's keep images as a separate "Always at end" block or auto-included.
    
    context = {
        'form': form,
        'image_formset': image_formset,
        'invoice': invoice,
        'title': f'Confirmation & Finalize - Invoice {invoice.tally_invoice_number or invoice.id}',
        'has_po': has_po,
        'has_email': has_email,
        'prev_url': prev_url,
        'packed_images_list': packed_images_list, 
        'current_step': 4,
        'progress_percentage': 90, # Not 100 yet
        'available_files': available_files
    }
    return render(request, 'clientdoc/confirmation_checklist.html', context)


def finalize_invoice_pdf(request, invoice_id):
    """Generates the final PDF based on user selected order."""
    invoice = get_object_or_404(SalesInvoice, id=invoice_id)
    confirmation = get_object_or_404(ConfirmationDocument, invoice=invoice)
    company_profile = OurCompanyProfile.objects.first()
    
    if request.method == 'POST':
        # Get order from POST
        # Valid separate IDs: invoice, dc, transport, po, email
        # We expect a comma separated string or list
        file_order_str = request.POST.get('file_order', 'invoice,dc,transport,po,email') 
        file_order = file_order_str.split(',')
        
        merger = PdfMerger()
        
        try:
            for file_type in file_order:
                if file_type == 'invoice':
                    # Check for Uploaded Custom Invoice first
                    if confirmation.uploaded_invoice:
                        try:
                            PdfReader(confirmation.uploaded_invoice.path)
                            merger.append(confirmation.uploaded_invoice.path)
                        except Exception:
                            # Fallback if corrupt? Or just fail. Fallback to generate.
                            invoice.calculate_total()
                            merger.append(generate_invoice_pdf(invoice, company_profile))
                    else:
                        invoice.calculate_total()
                        merger.append(generate_invoice_pdf(invoice, company_profile))
                
                elif file_type == 'dc':
                    if confirmation.uploaded_dc:
                        try:
                            PdfReader(confirmation.uploaded_dc.path)
                            merger.append(confirmation.uploaded_dc.path)
                        except Exception:
                             pass
                    elif hasattr(invoice, 'deliverychallan'):
                         merger.append(generate_dc_pdf(invoice, invoice.deliverychallan, company_profile))
                
                elif file_type == 'transport' and hasattr(invoice, 'transportcharges'):
                     merger.append(generate_transport_pdf(invoice, invoice.transportcharges, company_profile))
                
                elif file_type == 'po' and confirmation.po_file:
                    try:
                        PdfReader(confirmation.po_file.path)
                        merger.append(confirmation.po_file.path)
                    except Exception:
                        pass # Skip invalid
                
                elif file_type == 'email' and confirmation.approval_email_file:
                    try:
                        PdfReader(confirmation.approval_email_file.path)
                        merger.append(confirmation.approval_email_file.path)
                    except Exception:
                        pass

            # Always append images at the end
            images_pdf_buffer = generate_packed_images_pdf(confirmation)
            if images_pdf_buffer:
                merger.append(images_pdf_buffer)
            
            output = BytesIO()
            merger.write(output)
            merger.close()
            output.seek(0)
            
            # Save logic ...
            filename_suffix = invoice.tally_invoice_number or invoice.app_invoice_number or str(invoice.id)
            filename = f"confirmation_invoice_{filename_suffix}.pdf"
            
            # Preview vs Save
            path = os.path.join(settings.MEDIA_ROOT, 'confirmations', filename)
            os.makedirs(os.path.dirname(path), exist_ok=True) 

            with open(path, 'wb') as f:
                f.write(output.getvalue())
            
            confirmation.combined_pdf.name = f'confirmations/{filename}'
            confirmation.save()
            
            invoice.status = 'FIN'
            invoice.save()
            log_activity("Finalize Invoice", f"Finalized Invoice {invoice.tally_invoice_number or invoice.id}")
            
            messages.success(request, f'Document Bundle Generated Successfully!')
            return redirect('clientdoc:confirmation_list')

        except Exception as e:
            logger.error(f"Error finalizing PDF: {e}")
            messages.error(request, f"Error finalizing PDF: {e}")
            return redirect('clientdoc:create_confirmation', invoice_id=invoice_id)
            
    return redirect('clientdoc:create_confirmation', invoice_id=invoice_id)

# --- BULK UPLOAD VIEWS ---

def bulk_upload_page(request):
    """Page to upload excel and view history."""
    uploads = BulkInvoiceUpload.objects.order_by('-uploaded_at')
    
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        upload_type = request.POST.get('upload_type', 'invoice') # Default to invoice
        
        if not file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Please upload a valid Excel file.')
            return redirect('clientdoc:bulk_upload_page')
            
        upload_record = BulkInvoiceUpload.objects.create(file=file)
        upload_record.log = f"Type: {upload_type.title()}\n"
        upload_record.save()
        
        try:
            if upload_type == 'buyer':
                process_buyer_upload(upload_record)
            elif upload_type == 'item':
                process_item_upload(upload_record)
            elif upload_type == 'location':
                process_location_upload(upload_record)
            else:
                process_invoice_upload(upload_record)
                
            messages.success(request, f'{upload_type.title()} file uploaded and processed successfully.')
        except Exception as e:
            import traceback
            error_msg = f"Error processing file: {str(e)}\n{traceback.format_exc()}"
            upload_record.status = 'Failed'
            upload_record.log += error_msg
            upload_record.save()
            messages.error(request, 'Error processing file. Check logs.')
            
        return redirect('clientdoc:bulk_upload_page')
        
    return render(request, 'clientdoc/bulk_upload.html', {
        'uploads': uploads,
        'title': 'Bulk Data Upload'
    })

def download_sample_excel(request):
    """Generates a sample excel file based on type with formatting, optionally with data."""
    import datetime
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.comments import Comment
    
    upload_type = request.GET.get('type', 'invoice')
    do_export = request.GET.get('export') == 'true'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{upload_type.title()} {'Data' if do_export else 'Template'}"
    
    # Define Headers based on Type
    if upload_type == 'buyer':
        headers = ["Buyer Name*", "Address", "GSTIN", "State", "Phone", "Email"]
        widths = [30, 40, 20, 20, 20, 30]
        
    elif upload_type == 'item':
        headers = ["Item Name*", "Category", "Article/SKU", "Description", "Price*", "GST Rate (0.18)*", "HSN Code", "Unit (Nos)"]
        widths = [30, 20, 20, 40, 15, 15, 15, 15]
        
    elif upload_type == 'location':
        headers = ["Location Name*", "Site Code", "Address", "City", "State", "GSTIN", "Priority"]
        widths = [30, 15, 40, 20, 20, 20, 15]
        
    else: # Invoice
        headers = [
            'Buyer Name', 'Location Name', 'Item Name', 'Item Description', 'Quantity', 'Unit Rate', 
            'SGST', 'CGST', 'IGST', 'Transport Charges', 'Total Amount', 
            'Generate Invoice (Yes/No)', 'Generate PDF (Yes/No)', 
            'Tally Invoice No. (Identifier)', 'Invoce Date', 
            "Buyer's Order No.", "Buyer's Order Date (YYYY-MM-DD)", 
            'Dispatch Doc No.', 'Dispatched Through', 'Destination', 
            'Delivery Note', 'Delivery Note Date (YYYY-MM-DD)', 
            'Mode/Terms of Payment', 'Reference No. & Date', 'Other References', 
            'Terms of Delivery', 'Remarks', 'DC Notes', 'Transport Description', 
            'Doc-1 Invoice (Path)', 'Doc-2 DC (Path)', 'Doc-3 Buyer Po (Path)', 'Doc 4 Email approal (Path)', 
            'Doc-images-1', 'Doc-images-2', 'Doc-images-3', 'Doc-images-4', 'Doc-images-5'
        ]
        # Widths mostly uniform
        widths = [25] * len(headers)
        widths[0] = 30 # Buyer
        widths[1] = 30 # Location
        widths[2] = 30 # Item
        widths[3] = 40 # Description

    ws.append(headers)
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid") # Grey
    blue_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid") # Blue
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        
    # Blue First Column Header
    ws['A1'].fill = blue_fill
    
    # Add Comments for Guidance (Invoice Only)
    if upload_type == 'invoice':
        comments_map = {
            'A1': "Select Buyer from dropdown or ensure exact name match.",
            'B1': "Select Location (Ship To) from dropdown.",
            'C1': "Select Item. Description/Price auto-fill if left blank.",
            'J1': "Fill amount to auto-generate Transport Bill.",
            'L1': "Must be 'Yes' to process row.",
            'M1': "Must be 'Yes' to bundle PDFs.",
            'N1': "Unique ID. Leave blank to auto-generate (Tsol-XXXXX). Use same ID on multiple rows to group items.",
            'U1': "If filled, Delivery Challan (DC) is auto-created.",
            'V1': "If filled, Delivery Challan (DC) is auto-created.",
            'AB1': "Notes for DC. If filled, DC is auto-created.",
            'AD1': "Absolute file path (e.g. C:\\Docs\\Inv.pdf). Overrides auto-gen invoice.",
            'AG1': "Absolute file path for Approval Email PDF."
        }
        for cell_coord, note in comments_map.items():
            if cell_coord in ws:
                ws[cell_coord].comment = Comment(note, "System")

    # Set Widths
    for i, width in enumerate(widths, 1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # ---- EXPORT DATA LOGIC ----
    if do_export:
        if upload_type == 'buyer':
            for obj in Buyer.objects.all():
                ws.append([
                    obj.name, obj.address, obj.gstin, obj.state, obj.phone, obj.email
                ])
        elif upload_type == 'item':
            for obj in Item.objects.select_related('category').all():
                ws.append([
                    obj.name, 
                    obj.category.name if obj.category else "", 
                    obj.article_code, 
                    obj.description, 
                    obj.price, 
                    float(obj.gst_rate) if obj.gst_rate else 0.00,
                    obj.hsn_code, 
                    obj.unit
                ])
        elif upload_type == 'location':
            for obj in StoreLocation.objects.all():
                ws.append([
                    obj.name, obj.site_code, obj.address, obj.city, obj.state, obj.gstin, obj.priority
                ])
    
    # Invoice Specific Logic (Dropdowns etc - Only for Templates/Invoice)
    if upload_type == 'invoice':
        # Add Data and Validations
        data_ws = wb.create_sheet("Reference Data")
        data_ws.sheet_state = 'hidden' 
        
        buyers = list(Buyer.objects.values_list('name', flat=True))
        locations = list(StoreLocation.objects.values_list('name', flat=True))
        
        # Item Data for Auto-Fill (Name, Price, GST)
        items_qs = Item.objects.all().values_list('name', 'price', 'gst_rate')
        items = list(items_qs) # List of tuples
        
        for i, b in enumerate(buyers, 1): data_ws.cell(row=i, column=1, value=b)
        for i, l in enumerate(locations, 1): data_ws.cell(row=i, column=2, value=l)
        
        # Items in Cols 3, 4, 5 (C, D, E) (Reference Sheet)
        for i, (name, price, gst) in enumerate(items, 1): 
            data_ws.cell(row=i, column=3, value=name)
            data_ws.cell(row=i, column=4, value=price)
            data_ws.cell(row=i, column=5, value=gst)

        # Named Ranges for Robust Dropdowns
        from openpyxl.workbook.defined_name import DefinedName
        
        # Helper to safer add named range
        def create_named_range(name, sheet_title, range_ref):
            d = DefinedName(name, attr_text=f"'{sheet_title}'!{range_ref}")
            wb.defined_names.add(d)

        if buyers:
            create_named_range("BuyerList", "Reference Data", f"$A$1:$A${len(buyers)}")
        if locations:
            create_named_range("LocList", "Reference Data", f"$B$1:$B${len(locations)}")
        if items:
            create_named_range("ItemList", "Reference Data", f"$C$1:$C${len(items)}")
        
        def add_val(col, valid_formula):
             dv = DataValidation(type="list", formula1=valid_formula, allow_blank=True)
             ws.add_data_validation(dv)
             dv.add(f"{col}2:{col}500")

        if buyers: add_val('A', "=BuyerList")
        if locations: add_val('B', "=LocList")
        if items: add_val('C', "=ItemList")
        
        # VLOOKUP Formulas
        # Item Name is C. Description is D (User fills). Quantity is E. Unit Rate is F.
        # We want Unit Rate (F) to auto-fill from Reference Data D (Price) based on C (Item Name).
        # Reference Data: C=Name, D=Price, E=GST
        
        nrows = 500
        for r in range(2, nrows + 1):
             # Price VLOOKUP
             ws[f'F{r}'] = f"=IFERROR(VLOOKUP(C{r}, 'Reference Data'!$C$1:$E${len(items)+1}, 2, FALSE), \"\")"
             
        # Yes/No Dropdowns for L and M (Indices 11, 12)
        # 0=A, 1=B, 2=C, 3=D, 4=E, 5=F, 6=G, 7=H, 8=I, 9=J, 10=K
        # 11 = L (Gen Invoice)
        # 12 = M (Gen PDF)
        
        dv_yn = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
        ws.add_data_validation(dv_yn)
        dv_yn.add("L2:L500")
        ws.add_data_validation(dv_yn) 
        dv_yn.add("M2:M500")
        
        # Defaults
        ws['L2'] = "Yes"
        ws['M2'] = "Yes" 
        ws['W2'] = "30 Days" # Mode/Terms (Shifted: Old was V(21). Now 22(W))
        ws['Y2'] = "EMAIL Approval" # Other Ref (Old X(23). Now 24(Y))
    
    # Timestamped Filename
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")
    mode = "Export" if do_export else "Template"
    filename = f"Bulk_{upload_type.title()}_{mode}_{timestamp}.xlsx"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

# --- PROCESSORS ---
def process_buyer_upload(record):
    ws = openpyxl.load_workbook(record.file.path, data_only=True).active
    log = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or not row[0]: continue
        name = str(row[0]).strip()
        defaults = {
            'address': row[1] or "",
            'gstin': row[2] or "",
            'state': row[3] or "Karnataka",
            'phone': row[4] or "",
            'email': row[5] or ""
        }
        obj, created = Buyer.objects.update_or_create(name=name, defaults=defaults)
        log.append(f"Row {idx}: {'Created' if created else 'Updated'} Buyer '{name}'")
    
    record.log += "\n".join(log)
    record.status = 'Processed'
    record.save()

def process_item_upload(record):
    ws = openpyxl.load_workbook(record.file.path, data_only=True).active
    log = []
    from decimal import Decimal
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or not row[0]: continue
        name = str(row[0]).strip()
        
        # Category Logic
        cat_name = row[1]
        category = None
        if cat_name:
            category, _ = ItemCategory.objects.get_or_create(name=str(cat_name).strip())
            
        price = 0.00
        try: price = float(row[4]) if row[4] else 0.00
        except: pass
        
        gst = 0.18
        try: gst = float(row[5]) if row[5] else 0.18
        except: pass

        defaults = {
            'category': category,
            'article_code': row[2] or "",
            'description': row[3] or "",
            'price': Decimal(price),
            'gst_rate': Decimal(gst),
            'hsn_code': row[6] or "844311",
            'unit': row[7] or "Nos"
        }
        obj, created = Item.objects.update_or_create(name=name, defaults=defaults)
        log.append(f"Row {idx}: {'Created' if created else 'Updated'} Item '{name}'")
        
        record.log += "\n".join(log)
        record.status = 'Processed'
        record.save()

def process_location_upload(record):
    ws = openpyxl.load_workbook(record.file.path, data_only=True).active
    log = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or not row[0]: continue
        name = str(row[0]).strip()
        defaults = {
            'site_code': row[1] or "",
            'address': row[2] or "",
            'city': row[3] or "",
            'state': row[4] or "Karnataka",
            'gstin': row[5] or "",
            'priority': row[6] or ""
        }
        obj, created = StoreLocation.objects.update_or_create(name=name, defaults=defaults)
        log.append(f"Row {idx}: {'Created' if created else 'Updated'} Location '{name}'")
        
    record.log += "\n".join(log)
    record.status = 'Processed'
    record.save()

def process_invoice_upload(upload_record):
    """Parses Excel with support for Multiple Items per Invoice using Grouping - Updated Mapping & De-duplications"""
    file_path = upload_record.file.path
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    log = []
    created_count = 0
    updated_count = 0
    error_count = 0
    
    from datetime import datetime
    from decimal import Decimal
    import uuid
    from django.core.files import File
    import os
    
    def parse_date(date_val):
        if not date_val: return None
        if isinstance(date_val, datetime): return date_val
        try: return datetime.strptime(str(date_val).strip(), '%Y-%m-%d')
        except ValueError: return None 

    # --- 1. READ AND GROUP DATA ---
    grouped_rows = {} 
    
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row): continue
        
        def get_col(idx): return row[idx] if idx < len(row) else None
        
        # Mappings Updated (Inserted Description @ 3)
        # 0: Buyer, 1: Location, 2: Item, 3: DESC (NEW)
        # 4: Qty, 5: Unit Rate, 6: SGST, 7: CGST, 8: IGST, 9: Trans Charges, 10: Total
        # 11: Gen Inv, 12: Gen PDF
        # 13: Tally Inv
        # 14: Inv Date
        
        gen_invoice = get_col(11)
        if not gen_invoice or str(gen_invoice).strip().lower() != 'yes':
             log.append(f"Row {index}: Skipped (Generate != Yes)")
             continue

        location_name = get_col(1)
        item_name = get_col(2)
        qty = get_col(4)
        
        if not (location_name and item_name and qty):
             log.append(f"Row {index}: Skipped (Missing essential Item/Location data)")
             error_count += 1
             continue
             
        tally_no = str(get_col(13)).strip() if get_col(13) else None
        
        if tally_no:
            key = f"TALLY::{tally_no}"
        else:
            key = f"UNIQUE::{uuid.uuid4()}" 
            
        if key not in grouped_rows:
            grouped_rows[key] = []
        
        row_data = {
            'index': index,
            'buyer_name': get_col(0),
            'location_name': location_name,
            'item_name': item_name,
            'item_desc': get_col(3), # New Description
            'qty': qty,
            'unit_rate': get_col(5),
            'trans_charges': get_col(9),
            'gen_pdf': get_col(12),
            'tally_no': tally_no,
            'inv_date': parse_date(get_col(14)),
            'buyer_ord_no': get_col(15),
            'buyer_ord_date': parse_date(get_col(16)),
            'disp_doc_no': get_col(17),
            'disp_through': get_col(18),
            'dest': get_col(19),
            'del_note': get_col(20),
            'del_note_date': parse_date(get_col(21)),
            'pay_terms': get_col(22) or "30 Days",
            'ref_no': get_col(23),
            'other_ref': get_col(24) or "EMAIL Approval",
            'terms_del': get_col(25),
            'remark': get_col(26),
            'dc_notes': get_col(27),
            'trans_desc': get_col(28),
            # File Paths
            'doc_inv': get_col(29),
            'doc_dc': get_col(30),
            'doc_po': get_col(31),
            'doc_email': get_col(32),
            'doc_img_1': get_col(33),
            'doc_img_2': get_col(34),
            'doc_img_3': get_col(35),
            'doc_img_4': get_col(36),
            'doc_img_5': get_col(37),
        }
        grouped_rows[key].append(row_data)

    # --- 2. PROCESS GROUPS ---
    for key, rows in grouped_rows.items():
        first_row = rows[0]
        row_indices = [str(r['index']) for r in rows]
        indices_str = ", ".join(row_indices)
        
        try:
            with transaction.atomic():
                loc_obj = StoreLocation.objects.filter(name__iexact=str(first_row['location_name']).strip()).first()
                if not loc_obj:
                    log.append(f"Rows {indices_str}: Failed - Location '{first_row['location_name']}' not found")
                    error_count += 1
                    continue
                
                buyer_obj = None
                if first_row['buyer_name']:
                    buyer_obj = Buyer.objects.filter(name__iexact=str(first_row['buyer_name']).strip()).first()
                
                invoice = None
                is_update = False
                
                if first_row['tally_no']:
                     invoice = SalesInvoice.objects.filter(tally_invoice_number__iexact=first_row['tally_no']).first()
                     if invoice: is_update = True
                
                header_data = {
                    'buyer': buyer_obj,
                    'location': loc_obj,
                    'tally_invoice_number': first_row['tally_no'],
                    'buyers_order_no': first_row['buyer_ord_no'],
                    'buyers_order_date': first_row['buyer_ord_date'] or datetime.now(),
                    'dispatch_doc_no': first_row['disp_doc_no'],
                    'dispatched_through': first_row['disp_through'],
                    'destination': first_row['dest'],
                    'delivery_note': first_row['del_note'],
                    'delivery_note_date': first_row['del_note_date'] or datetime.now(),
                    'mode_terms_payment': first_row['pay_terms'],
                    'reference_no_date': first_row['ref_no'],
                    'other_references': first_row['other_ref'],
                    'terms_of_delivery': first_row['terms_del'],
                    'remark': first_row['remark'],
                }
                
                if first_row['inv_date']: header_data['date'] = first_row['inv_date']

                if is_update and invoice:
                     for k, v in header_data.items():
                         if v is not None: setattr(invoice, k, v)
                     invoice.save()
                     log.append(f"Rows {indices_str}: Updated Invoice {invoice.app_invoice_number or invoice.id}")
                     updated_count += 1
                else:
                    if 'date' not in header_data: header_data['date'] = datetime.now()
                    header_data['status'] = 'DRF'
                    invoice = SalesInvoice.objects.create(**header_data)
                    log.append(f"Rows {indices_str}: Created Invoice #{invoice.id}")
                    created_count += 1
                    
                # --- PROCESS ITEMS (Iterate ALL rows in group) ---
                for r in rows:
                    item_obj = Item.objects.filter(name__iexact=str(r['item_name']).strip()).first()
                    if not item_obj:
                         log.append(f"Row {r['index']}: Warning - Item '{r['item_name']}' not found. Skipped.")
                         continue
                    try: q = int(r['qty'])
                    except: q = 1
                    
                    price = item_obj.price
                    if r['unit_rate']:
                        try: price = Decimal(str(r['unit_rate']).strip())
                        except: pass
                    
                    # Prevent Duplicates and Fix "Returned more than one" error
                    # If multiple items exist (from previous bad uploads), delete them first.
                    existing_dupes = InvoiceItem.objects.filter(invoice=invoice, item=item_obj)
                    if existing_dupes.count() > 1:
                        existing_dupes.delete()

                    # Update or Create based on Item
                    InvoiceItem.objects.update_or_create(
                        invoice=invoice,
                        item=item_obj,
                        defaults={
                            'quantity': q,
                            'quantity_billed': q,
                            'quantity_shipped': q,
                            'price': price,
                            'gst_rate': item_obj.gst_rate,
                            'description': r['item_desc'] 
                        }
                    )
                
                # Create DC if Notes OR Delivery Note details are present
                if first_row['dc_notes'] or first_row['del_note'] or first_row['del_note_date']:
                    dc, _ = DeliveryChallan.objects.get_or_create(invoice=invoice)
                    if first_row['dc_notes']: 
                        dc.notes = first_row['dc_notes']
                    dc.save()
                    if invoice.status == 'DRF': invoice.status = 'DC'
                    
                if first_row['trans_charges']:
                     try:
                         amt = Decimal(str(first_row['trans_charges']).strip()) 
                         trp, _ = TransportCharges.objects.get_or_create(invoice=invoice)
                         trp.charges = amt
                         trp.description = first_row['trans_desc']
                         trp.save()
                         # Force invoice to be aware if needed or just status update
                         if invoice.status in ['DRF', 'DC']: invoice.status = 'TRP'
                     except Exception as e:
                         log.append(f"Row {first_row['index']}: Warning - Invalid Transport Charge ({e})")
                
                invoice.save()
                
                # CRITICAL: Calculate total AFTER adding transport charges so Tax Matrix includes them
                # refresh_from_db isn't strictly needed inside atomic for related objects unless cached, 
                # but let's be safe for calculate logic.
                if hasattr(invoice, 'transportcharges'): invoice.transportcharges.refresh_from_db()
                invoice.calculate_total() 
                
                # --- FILE UPLOADS ---
                # Fix: Check all_objects to handle soft-deleted records to prevent UNIQUE constraint error
                conf = ConfirmationDocument.all_objects.filter(invoice=invoice).first()
                if conf:
                    if conf.is_deleted:
                        conf.restore()
                else:
                    conf = ConfirmationDocument.objects.create(invoice=invoice)
                
                def save_file_from_path(path_val, target_field):
                    if path_val:
                         path_val = str(path_val).strip() # Clean path
                         if os.path.exists(path_val):
                             try:
                                 with open(path_val, 'rb') as f:
                                     fname = os.path.basename(path_val)
                                     target_field.save(fname, File(f), save=True)
                             except Exception as fe:
                                 log.append(f" Failed to load file {path_val}: {fe}")
                         else:
                             log.append(f" File not found: {path_val}")
                
                save_file_from_path(first_row['doc_po'], conf.po_file)
                save_file_from_path(first_row['doc_email'], conf.approval_email_file)
                save_file_from_path(first_row['doc_inv'], conf.uploaded_invoice)
                save_file_from_path(first_row['doc_dc'], conf.uploaded_dc)
                
                # --- PACKED IMAGES (Iterate 5 slots) ---
                img_slots = [first_row[f'doc_img_{i}'] for i in range(1, 6)]
                for img_path in img_slots:
                    if img_path:
                        img_path = str(img_path).strip()
                        if os.path.exists(img_path):
                            try:
                                with open(img_path, 'rb') as f:
                                    pi = PackedImage(confirmation=conf)
                                    pi.image.save(os.path.basename(img_path), File(f), save=True)
                            except Exception as ie:
                               log.append(f" Failed to load image {img_path}: {ie}")
                        else:
                            log.append(f" Image not found: {img_path}")

                # --- PDF GENERATION ---
                should_gen_pdf = any(str(r['gen_pdf']).strip().lower() == 'yes' for r in rows if r['gen_pdf'])
                if should_gen_pdf:
                    try:
                        company_profile = OurCompanyProfile.objects.first()
                        merger = PdfMerger()
                        
                        # 1. Tax Invoice
                        if conf.uploaded_invoice:
                            try:
                                PdfReader(conf.uploaded_invoice.path)
                                merger.append(conf.uploaded_invoice.path)
                            except: pass 
                        else:
                            # Re-Calculate to be 100% sure before PDF Gen
                            invoice.calculate_total()
                            merger.append(generate_invoice_pdf(invoice, company_profile))
                            
                        # 2. Delivery Note (DC)
                        if conf.uploaded_dc:
                            try:
                                PdfReader(conf.uploaded_dc.path)
                                merger.append(conf.uploaded_dc.path)
                            except: pass
                        elif hasattr(invoice, 'deliverychallan'):
                            merger.append(generate_dc_pdf(invoice, invoice.deliverychallan, company_profile))
                            
                        # 3. Transport Charges
                        if hasattr(invoice, 'transportcharges'):
                             # Ensure the relation is accessible
                             merger.append(generate_transport_pdf(invoice, invoice.transportcharges, company_profile))
                        
                        # 4. Email Approval / Buyer PO
                        # User requested "Email Approval / Buyer PO". We'll append Email first, then PO if desired, or prioritized.
                        # Assuming they map to the same conceptual "Approval" step. 
                        conf.refresh_from_db() 
                        
                        if conf.approval_email_file:
                             try: merger.append(conf.approval_email_file.path)
                             except: pass
                             
                        if conf.po_file:
                             try: merger.append(conf.po_file.path)
                             except: pass

                        # 5. Images (1-5)
                        images_pdf_buffer = generate_packed_images_pdf(conf)
                        if images_pdf_buffer:
                            merger.append(images_pdf_buffer)

                        output = BytesIO()
                        merger.write(output)
                        merger.close()
                        output.seek(0)
                        
                        suffix = invoice.tally_invoice_number or invoice.app_invoice_number or str(invoice.id)
                        filename = f"confirmation_invoice_{suffix}.pdf"
                        
                        from django.core.files.base import ContentFile
                        conf.combined_pdf.save(filename, ContentFile(output.getvalue()), save=True)
                        invoice.status = 'FIN'
                        invoice.save()
                        log.append(f" Invoice #{invoice.id}: PDF Generated (Bundled)")
                    except Exception as pdf_err:
                        logger.error(f"Bulk PDF Error: {pdf_err}")
                        log.append(f" Invoice #{invoice.id}: PDF Failed ({str(pdf_err)})")

        except Exception as e:
            log.append(f"Rows {indices_str}: Group Error - {str(e)}")
            error_count += 1
            import traceback
            logger.error(traceback.format_exc())

    upload_record.log = "\n".join(log)
    upload_record.status = 'Processed'
    upload_record.save()
    
    return redirect('clientdoc:dashboard')

def create_buyer(request):
    if request.method == 'POST':
        form = BuyerForm(request.POST) 
        if form.is_valid():
            form.save()
            log_activity("Create Buyer", f"Created Buyer {form.instance.name}")
            messages.success(request, 'Buyer created successfully.')
            return redirect('clientdoc:dashboard')
    else:
        form = BuyerForm()
    return render(request, 'clientdoc/form.html', {'form': form, 'title': 'Create Buyer'})

def buyer_list(request):
    search_fields = ['name', 'address', 'gstin', 'state']
    buyers = get_filtered_queryset(Buyer, request, search_fields)
    
    paginator = Paginator(buyers, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'clientdoc/buyer_list.html', {
        'page_obj': page_obj, 
        'title': 'Buyer List',
        'list_type': 'buyer'
    })

def edit_buyer(request, pk):
    buyer = get_object_or_404(Buyer, pk=pk)
    if request.method == 'POST':
        form = BuyerForm(request.POST, instance=buyer)
        if form.is_valid():
            form.save()
            log_activity("Edit Buyer", f"Updated Buyer {buyer.name}")
            messages.success(request, 'Buyer updated successfully.')
            return redirect('clientdoc:buyer_list')
    else:
        form = BuyerForm(instance=buyer)
    return render(request, 'clientdoc/form.html', {'form': form, 'title': 'Edit Buyer'})

def buyer_detail(request, pk):
    buyer = get_object_or_404(Buyer, pk=pk)
    return render(request, 'clientdoc/buyer_detail.html', {'buyer': buyer})


def delete_packed_image(request, image_id):
    """Handles the deletion of a specific packed image, ensuring file removal."""
    image = get_object_or_404(PackedImage, id=image_id)
    invoice_id = image.confirmation.invoice.id
    
    if request.method == 'POST':
        if image.image:
            image.image.delete(save=False) 
        
        image.delete()
        messages.success(request, 'Image successfully removed.')
    else:
        messages.error(request, 'Invalid request method.')
        
    return redirect('clientdoc:create_confirmation', invoice_id=invoice_id)

def print_invoice(request, invoice_id):
    """Renders the print-friendly invoice template."""
    invoice = get_object_or_404(SalesInvoice, id=invoice_id)
    company_profile = OurCompanyProfile.objects.first()
    
    # Ensure totals are calculated
    invoice.calculate_gst_totals()
    
    display_invoice_number = invoice.tally_invoice_number if invoice.tally_invoice_number else invoice.app_invoice_number
    
    # Determine IGST vs CGST/SGST based on model's calculated fields
    # Logic: If igst_total > 0, it's Inter-state. Or check place_of_supply vs company state.
    # However, model stores totals now.
    
    comp_state_code = company_profile.state_code if company_profile else '29'
    # Fallback to model POS if set, else Location state
    pos_code = invoice.place_of_supply if invoice.place_of_supply else (invoice.location.state_code if invoice.location else '29')
    
    is_igst = (pos_code != comp_state_code)
    
    # Re-sum taxable for display if needed, or rely on grand total - tax? 
    # Better to sum line items for the "Taxable Value" column/row in template.
    taxable_val = sum(item.taxable_value for item in invoice.invoiceitem_set.all())
    
    if hasattr(invoice, 'transportcharges') and invoice.transportcharges and invoice.transportcharges.charges > 0:
        taxable_val += invoice.transportcharges.charges

    return render(request, 'clientdoc/invoice_print_template.html', {
        'invoice': invoice,
        'company': company_profile,
        'display_invoice_number': display_invoice_number,
        'taxable_val': taxable_val,
        'tax_amt': (invoice.cgst_total + invoice.sgst_total + invoice.igst_total),
        'cgst_amt': invoice.cgst_total,
        'sgst_amt': invoice.sgst_total,
        'igst_amt': invoice.igst_total,
        'is_igst': is_igst,
    })

def print_dc(request, invoice_id):
    """Renders the print-friendly Delivery Challan template."""
    invoice = get_object_or_404(SalesInvoice, id=invoice_id)
    # Get the associated Delivery Challan
    dc = get_object_or_404(DeliveryChallan, invoice=invoice)
    company_profile = OurCompanyProfile.objects.first()
    
    # Calculate total quantity
    total_qty = sum(item.quantity for item in invoice.invoiceitem_set.all())
    display_invoice_number = invoice.tally_invoice_number if invoice.tally_invoice_number else invoice.app_invoice_number
    
    return render(request, 'clientdoc/dc_print_template.html', {
        'invoice': invoice,
        'dc': dc,
        'company': company_profile,
        'total_qty': total_qty,
        'display_invoice_number': display_invoice_number
    })

def print_transport(request, invoice_id):
    """Renders the print-friendly Transport Charges template."""
    invoice = get_object_or_404(SalesInvoice, id=invoice_id)
    # Get the associated Transport Charges
    transport = get_object_or_404(TransportCharges, invoice=invoice)
    company_profile = OurCompanyProfile.objects.first()
    
    display_invoice_number = invoice.tally_invoice_number if invoice.tally_invoice_number else invoice.app_invoice_number
    
    return render(request, 'clientdoc/transport_print_template.html', {
        'invoice': invoice,
        'transport': transport,
        'company': company_profile,
        'display_invoice_number': display_invoice_number
    })
def project_guide(request):
    """Serves the Project Guide PDF."""
    import os
    from django.conf import settings
    from django.http import HttpResponse, Http404

    file_path = os.path.join(settings.BASE_DIR, 'Project guide', 'Project Guide.pdf')
    if os.path.exists(file_path):
        with open(file_path, 'rb') as pdf:
            response = HttpResponse(pdf.read(), content_type='application/pdf')
            response['Content-Disposition'] = 'inline; filename="Project Guide.pdf"'
            return response
    else:
        raise Http404("Project Guide not found")
